import pandas as pd
import streamlit as st
from fpdf import FPDF

from datetime import datetime
import io
import os
import base64
import openpyxl


# Configuración de la página
st.set_page_config(
    page_title="Sistema de Análisis de Alelos",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado para diseño atractivo
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #2E86AB;
        text-align: center;
        margin-bottom: 2rem;
        font-weight: bold;
        background: linear-gradient(90deg, #2E86AB, #A23B72);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .sub-header {
        font-size: 1.4rem;
        color: #2E86AB;
        margin-bottom: 1rem;
        border-bottom: 2px solid #2E86AB;
        padding-bottom: 0.5rem;
    }
    .patient-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 25px;
        border-radius: 15px;
        color: white;
        margin-bottom: 20px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    .result-box {
        background: #f8f9fa;
        padding: 20px;
        border-radius: 10px;
        border-left: 5px solid #2E86AB;
        margin: 10px 0;
    }
    .stButton>button {
        background: linear-gradient(135deg, #2E86AB, #A23B72);
        color: white;
        border-radius: 8px;
        padding: 12px 24px;
        border: none;
        font-weight: bold;
        transition: all 0.3s ease;
    }
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.2);
    }
    .upload-box {
        border: 2px dashed #2E86AB;
        border-radius: 10px;
        padding: 20px;
        text-align: center;
        background-color: #f8f9fa;
        margin: 10px 0;
    }
</style>
""", unsafe_allow_html=True)


def create_pdf(paciente_data, datos_geneticos, paciente_codigo):
    pdf = FPDF()
    pdf.add_page()
    
    # Header
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'INFORME DE FARMACOGENÉTICA', 0, 1, 'C')
    pdf.set_font('Arial', 'I', 10)
    pdf.cell(0, 10, f'Fecha: {datetime.now().strftime("%d/%m/%Y")}', 0, 1, 'C')
    pdf.ln(10)
    
    # Datos del paciente
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'DATOS DEL PACIENTE', 0, 1)
    pdf.ln(5)
    
    pdf.set_font('Arial', '', 12)
    info_lines = [
        f"Nombre: {paciente_data.get('nombre', 'No especificado')}",
        f"Edad: {paciente_data.get('edad', 'No especificado')}",
        f"Sexo: {paciente_data.get('sexo', 'No especificado')}",
        f"ID Paciente: {paciente_data.get('id_paciente', paciente_codigo)}",
        f"Médico: {paciente_data.get('medico', 'No especificado')}",
        f"Fecha de Nacimiento: {paciente_data.get('fecha_nacimiento', 'No especificada')}"
    ]
    
    for line in info_lines:
        pdf.cell(0, 8, line, 0, 1)
    
    pdf.ln(10)
    
    # Resultados genéticos
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'RESULTADOS GENÉTICOS', 0, 1)
    pdf.ln(5)
    
    if datos_geneticos:
        for gen, info in datos_geneticos.items():
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 8, f"Gen: {gen}", 0, 1)
            
            pdf.set_font('Arial', '', 10)
            pdf.cell(0, 6, f"Genotipo: {info[0]}", 0, 1)
            pdf.cell(0, 6, f"Score: {info[1]}", 0, 1)
            pdf.cell(0, 6, f"Fenotipo: {info[2]}", 0, 1)
            
            # Recomendación clínica (puede ser larga, usar multi_cell)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 6, "Recomendación clínica:", 0, 1)
            pdf.set_font('Arial', '', 9)
            
            # Dividir texto largo en líneas
            recomendacion = info[3]
            # Ajustar el ancho de línea para que quepa en la página
            pdf.multi_cell(0, 5, recomendacion)
            
            pdf.ln(5)
    else:
        pdf.set_font('Arial', 'I', 12)
        pdf.cell(0, 10, 'No hay datos genéticos disponibles', 0, 1)
    
    # Información adicional si existe
    if paciente_data.get('observaciones'):
        pdf.ln(5)
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 8, "Observaciones clínicas:", 0, 1)
        pdf.set_font('Arial', '', 10)
        pdf.multi_cell(0, 6, paciente_data.get('observaciones', ''))
    
    return pdf

def lectura_csv(path):
    # Lectura del archivo csv con los datos de cada paciente.
    df = pd.read_csv(path, sep=';')
    # Los datos se pasan a un diccionario
    datos_dict = {df["Sample/Assay"][i]: {df.columns[j]: df[df.columns[j]][i] for j in range(1, len(df.columns))} for i in range(len(df))}
    # Tabla con las variantes y sus mutaciones asociadas
    tabla = {"CYP2D6": {"3": ["_"], "4": ["T"], "6": [ "_"], "7": ["G"],
                    "8": ["A"], "9": ["_"], "10*4": ["A"], "10": ["G"],
                    "12": ["T"], "14": ["T"], "15": ["_"], "17": ["A"],
                    "19": ["_"], "29": ["A"], "41": ["T"], "56B": ["A"],
                    "59": ["T"]},
         "UGT1A1": {"80": ["T"]},
         "DPYD": {"2A": [ "G","T" ], "13": ["C", "T"], "HapB3": ["T"], "D949V": ["A"]}}

    dict_pacientes = {}


    for i in datos_dict:
        dict_pacientes[i] = {}
        for j in datos_dict[i]:
            # Separacion de nombre de gen y variantes
            if "*" in j:
                separador = "*"
            elif "_" in j:
                separador = "_"
            snp = j.split(separador, 1)


            if snp[0] not in dict_pacientes[i]:
                dict_pacientes[i][snp[0]] = []

            # Manejo de Indeterminados
            if datos_dict[i][j] == "UND":
                dict_pacientes[i][snp[0]].append(("*1", "*1"))
                continue
            # Determinar genotipo 
            variante = []
            if datos_dict[i][j].split("/")[0] in tabla[snp[0]][snp[1]]:
                variante.append(f"{separador}{snp[1]}")
            else:
                variante.append(f"*1")

            if datos_dict[i][j].split("/")[1] in tabla[snp[0]][snp[1]]:
                variante.append(f"{separador}{snp[1]}")
            else:
                variante.append(f"*1")

            dict_pacientes[i][snp[0]].append(tuple(variante))

    # Eliminar duplicados
    for i in dict_pacientes:
        for j in dict_pacientes[i]:
            dict_pacientes[i][j] = list(set(dict_pacientes[i][j]))

    return dict_pacientes

def determinar_genotipo_definitivo(datos_pacientes):
    """
    Determina el genotipo definitivo para cada gen de cada paciente
    """
    resultados = {}
    
    for paciente, genes in datos_pacientes.items():
        resultados[paciente] = {}
        
        for gen, alelos in genes.items():
            # Recoger todos los alelos únicos, excluyendo *1 cuando hay otros
            alelos_maternos_unicos = set()
            alelos_paternos_unicos = set()

            for alelo_materno, alelo_paterno in alelos:
                # Si ambos son *1, es el caso base (sin mutaciones)
                if alelo_materno == '*1' and alelo_paterno == '*1':
                    continue
                
                # Añadir alelos no silvestres
                
                if alelo_materno != '*1':
                    alelos_maternos_unicos.add(alelo_materno)
                if alelo_paterno != '*1':
                    alelos_paternos_unicos.add(alelo_paterno)

            if gen == "CYP2D6":
                # Manejo de Variante *10*4 Materno
                if "*10*4" in alelos_maternos_unicos and "*10" in alelos_maternos_unicos:
                    alelos_maternos_unicos.remove("*10*4")
                    if "*4" in alelos_maternos_unicos:
                        alelos_maternos_unicos.remove("*10")
                # Manejo de Variante *10*4 Paterno 
                if "*10*4" in alelos_paternos_unicos and "*10" in alelos_paternos_unicos:
                    alelos_paternos_unicos.remove("*10*4")
                    if "*4" in alelos_paternos_unicos:
                        alelos_paternos_unicos.remove("*10")
                #Manejo si estan en alelos distintos
                if "*10*4" in alelos_paternos_unicos and "*10" in alelos_maternos_unicos:
                    alelos_paternos_unicos.remove("*10*4")
                    if "*4" in alelos_maternos_unicos:
                        alelos_maternos_unicos.remove("*10")
                if "*10*4" in alelos_maternos_unicos and "*10" in alelos_paternos_unicos:
                    alelos_maternos_unicos.remove("*10*4")
                    if "*4" in alelos_paternos_unicos:
                        alelos_paternos_unicos.remove("*10")


            # Manejo de variante *80 como *28
            if gen == "UGT1A1":
                if alelos_maternos_unicos:
                    alelos_maternos_unicos = list(alelos_maternos_unicos)
                    alelos_maternos_unicos[0] = "*28"
                    alelos_maternos_unicos = set(alelos_maternos_unicos)
                if alelos_paternos_unicos:
                    alelos_paternos_unicos = list(alelos_paternos_unicos)
                    alelos_paternos_unicos[0] = "*28"
                    alelos_paternos_unicos = set(alelos_paternos_unicos)

            # Si no hay mutaciones, el genotipo es *1/*1
            if not alelos_maternos_unicos and not alelos_paternos_unicos:
                resultados[paciente][gen] = ('*1', '*1')
            # Si hay una mutación, es heterocigoto *1/mutación
            elif alelos_maternos_unicos and not alelos_paternos_unicos:
                mutacion = list(alelos_maternos_unicos)[0]
                resultados[paciente][gen] = ('*1', mutacion)
            # Si hay dos mutaciones, es heterocigoto mutación1/mutación2
            elif not alelos_maternos_unicos and alelos_paternos_unicos:
                mutacion = list(alelos_paternos_unicos)[0]
                resultados[paciente][gen] = ('*1', mutacion)
            elif alelos_maternos_unicos and alelos_paternos_unicos:

                mutacion_materna = list(alelos_maternos_unicos)[0]
                mutacion_paterna = list(alelos_paternos_unicos)[0]
                resultados[paciente][gen] = (mutacion_materna, mutacion_paterna)

    return resultados

# Función para formatear el resultado como string
def formatear_genotipos(resultados):
    formateados = {}
    for paciente, genes in resultados.items():
        formateados[paciente] = {}
        for gen, alelos in genes.items():
            formateados[paciente][gen] = f"{alelos[0]}/{alelos[1]}"
    return formateados

df = pd.read_excel("CYP2D6_Diplotype_Phenotype_Table.xlsx")
diccionario_CYP2D6 = dict(zip(df.iloc[:, 0], zip(df.iloc[:, 1], df.iloc[:, 2])))
def fenotipo(genotipo):
    Sol = {}
    diccionario = genotipo
    for nombre in diccionario:
        Sol[nombre] = {}
        for gen in diccionario[nombre]:
            alelos = diccionario[nombre][gen].split('/')   
            if gen == 'DPYD' or gen == 'UGT1A1':
                if alelos[0] == "*1" and alelos[1] == "*1":       
                    Sol[nombre][gen] = [f"{alelos[0]}/{alelos[1]}", 2.0, 'Normal Metabolizer']
                elif alelos[0] != "*1" and alelos[1] != "*1":
                    Sol[nombre][gen] = [f"{alelos[0]}/{alelos[1]}", 0.0, 'Poor Metabolizer']
                else:
                    Sol[nombre][gen] = [f"{alelos[0]}/{alelos[1]}", 1.0,'Intermediate Metabolizer']
            else:
                Sol[nombre][gen] = [diccionario[nombre][gen]]
                Sol[nombre][gen].append(diccionario_CYP2D6[diccionario[nombre][gen]][0])
                Sol[nombre][gen].append(diccionario_CYP2D6[diccionario[nombre][gen]][1])
    return Sol


def recomendacionClinica(fenotipo):
    import json # Importa la biblioteca JSON para trabajar con datos JSON.
    import requests # Importa la biblioteca Requests para realizar solicitudes HTTP.
    resultado = fenotipo # Inicializa una lista vacía para almacenar los resultados.
    for paciente in fenotipo:
        for gen in fenotipo[paciente]:
            if gen == "CYP2D6":
                lookupkey= [gen, str(fenotipo[paciente][gen][1])] # Obtiene la clave de búsqueda del fenotipo.
                ID_Farmaco = "RxNorm:10324"
                url='https://api.cpicpgx.org/v1/recommendation?select=drug(name),guideline(name),*&drugid=eq.'+ID_Farmaco+'&lookupkey=cs.{\"'+lookupkey[0]+'":"'+lookupkey[1]+'"}' 
            elif gen == "DPYD":
                lookupkey= [gen, str(fenotipo[paciente][gen][1])] # Obtiene la clave de búsqueda del fenotipo.
                ID_Farmaco = "RxNorm:51499"
                url='https://api.cpicpgx.org/v1/recommendation?select=drug(name),guideline(name),*&drugid=eq.'+ID_Farmaco+'&lookupkey=cs.{\"'+lookupkey[0]+'":"'+lookupkey[1]+'"}' 
                if float(lookupkey[1])==2.0:
                    resultado[paciente][gen].append("Based on genotype, there is no indication to change dose or therapy. Use label-recommended dosage and administration.")
                elif float(lookupkey[1])>=1.0:
                    resultado[paciente][gen].append("Reduce starting dose by 50% followed by titration of dose based on toxicity or therapeutic drug monitoring (if available). Patients with the c.[2846A>T];[2846A>T] genotype may require >50% reduction in starting dose.")
                elif float(lookupkey[1])==0.5:
                    resultado[paciente][gen].append("Avoid use of 5- fluorouracil or 5-fluorouracil prodrug-based regimens. In the event, based on clinical advice, alternative agents are not considered a suitable therapeutic option, 5-fluorouracil should be administered at a strongly reduced dose with early therapeutic drug monitoring.")
                elif float(lookupkey[1])== 0.0:
                    resultado[paciente][gen].append("Avoid use of 5-fluorouracil or 5-fluorouracil prodrug-based regimens.")
            elif gen == "UGT1A1":
                lookupkey= [gen, str(fenotipo[paciente][gen][1])] # Obtiene la clave de búsqueda del fenotipo.
                ID_Farmaco = "RxNorm:51499"
                url='https://api.cpicpgx.org/v1/recommendation?select=drug(name),guideline(name),*&drugid=eq.'+ID_Farmaco+'&lookupkey=cs.{\"'+lookupkey[0]+'":"'+lookupkey[1]+'"}' 
                if fenotipo[paciente][gen][0] == "*1/*1":
                    resultado[paciente][gen].append("The guideline does not provide a recommendation for irinotecan in normal metabolizers.")
                elif fenotipo[paciente][gen][0] == "*1/*28":
                    resultado[paciente][gen].append("NO action is needed for this gene-drug interaction.")
                elif fenotipo[paciente][gen][0] == "*28/*28":
                    resultado[paciente][gen].append("Start with 70% of the normal dose If the patient tolerates this initial dose, the dose can be increased, guided by the neutrophil count.")
            response = requests.get(url) # Realiza una solicitud GET a la API.
            json_obtenido = response.json() # Convierte la respuesta JSON en un objeto Python.
            datos=json_obtenido # Asigna los datos JSON a la variable 'datos'.
            if len(datos) != 0: # Verifica si se encontraron recomendaciones.
                resultado[paciente][gen].append(datos[0]['drugrecommendation'].encode('latin-1','ignore').decode('latin-1')) # Agrega la recomendación del fármaco a la lista, decodificando caracteres especiales.
    return resultado # Devuelve la lista con los resultados.


def main():
    # Header principal
    st.markdown('<div class="main-header">🧬 SISTEMA DE ANÁLISIS DE ALELOS</div>', unsafe_allow_html=True)
    
    # Sidebar para navegación
    with st.sidebar:
        st.markdown("### 📊 Navegación")
        page = st.radio("Selecciona una sección:", 
                       ["📤 Cargar Alelos","📝 Datos del Paciente", "📄 Generar Reporte"])
        
        st.markdown("---")
        st.markdown("### 💡 Información")
        st.info("""
        Esta aplicación permite:
        - Subir archivos CSV con alelos
        - Rellenar los datos de los pacientes
        - Generar reportes en PDF
        """)
    
    # Inicializar session state
    if 'paciente_data' not in st.session_state:
        st.session_state.paciente_data = None
    if 'alelos_df' not in st.session_state:
        st.session_state.alelos_df = None
    
    
    # Sección 1: Cargar Alelos
    if page == "📤 Cargar Alelos":
        st.markdown('<div class="sub-header">📤 CARGA DE ARCHIVOS CSV</div>', unsafe_allow_html=True)
        
        st.markdown('<div class="upload-box">', unsafe_allow_html=True)
        uploaded_file = st.file_uploader("Sube tu archivo CSV con los alelos de cada paciente", type=['csv'])
        st.markdown('</div>', unsafe_allow_html=True)
        
        if uploaded_file is not None:
            try:
                # Leer el archivo CSV
                dict_pacientes = lectura_csv(uploaded_file)
                st.session_state.pacientes_data = dict_pacientes
                
                st.success(f"✅ Archivo cargado correctamente! ({len(dict_pacientes)} pacientes)")
                
                st.markdown("### Procesando datos...")

                resultados = determinar_genotipo_definitivo(dict_pacientes)
                resultados_formateados = formatear_genotipos(resultados)
                fenotipo_resultado = fenotipo(resultados_formateados)
                resultado_final = recomendacionClinica(fenotipo_resultado)

                st.session_state.resultado = resultado_final

                st.success(f"Datos procesados! Por favor, pasa a la siguiente sección.")
                    
            except Exception as e:
                st.error(f"❌ Error al cargar el archivo: {str(e)}")
        
        # Ejemplo de formato esperado
        with st.expander("📝 ¿Qué formato debe tener el CSV?"):
            st.markdown("""
            **Ejemplo de estructura esperada:**
            ```
            Paciente;gen1alelo1;gen1alelo2;gen2alelo1;gen2alelo2
            DPD900;  T/T;       C/G;       T/C;       C/A
            DPD901;  T/T;       G/G;       C/C;       C/C
            DPD902;  T/C;       C/G;       C/T;       A/C
            ```
            Los genes que esta herramienta son capaces de procesar son 'CYP2D6', 'DPYD' y 'UGT1A1'.
            """)


    # Seccion 2: Datos de los pacientes

    elif page == "📝 Datos del Paciente":
        st.markdown('<div class="sub-header">📝 INFORMACIÓN DE LOS PACIENTES</div>', unsafe_allow_html=True)
        
        # Inicializar el almacenamiento de pacientes en session_state si no existe
        if 'pacientes_data' not in st.session_state:
            st.session_state.pacientes_data = {}
        
        # Selector de paciente
        pacientes_disponibles = list(st.session_state.resultado.keys())
        paciente_seleccionado = st.selectbox("Selecciona el paciente *", pacientes_disponibles)
        
        st.markdown(f"### Editando datos para: **{paciente_seleccionado}**")
        
        # Mostrar información genética del paciente seleccionado
        with st.expander("🧬 Ver información genética del paciente"):
            if paciente_seleccionado in st.session_state.resultado:
                datos_geneticos = st.session_state.resultado[paciente_seleccionado]
                for gen, info in datos_geneticos.items():
                    st.markdown(f"**{gen}:**")
                    st.markdown(f"- Genotipo: {info[0]}")
                    st.markdown(f"- Score: {info[1]}")
                    st.markdown(f"- Fenotipo: {info[2]}")
                    st.markdown(f"- Recomendación: {info[3]}")
        
        with st.form(f"patient_form_{paciente_seleccionado}"):
            # Cargar datos existentes si los hay, sino valores por defecto/vacíos
            datos_existentes = st.session_state.pacientes_data.get(paciente_seleccionado, {})
            
            col1, col2 = st.columns(2)
            
            with col1:
                nombre = st.text_input(
                    "Nombre completo", 
                    value=datos_existentes.get('nombre', ''),
                    placeholder="Ej: Juan Pérez García"
                )
                edad = st.number_input(
                    "Edad", 
                    min_value=0, 
                    max_value=120, 
                    value=datos_existentes.get('edad', 0),
                    placeholder="Ej: 45"
                )
                sexo = st.selectbox(
                    "Sexo", 
                    ["", "Masculino", "Femenino", "Otro"],
                    index=0 if 'sexo' not in datos_existentes else 
                        ["", "Masculino", "Femenino", "Otro"].index(datos_existentes.get('sexo', ''))
                )
            
            with col2:
                id_paciente = st.text_input(
                    "ID del Paciente", 
                    value=datos_existentes.get('id_paciente', paciente_seleccionado),
                    placeholder="Ej: PAC-001"
                )
                medico = st.text_input(
                    "Médico solicitante", 
                    value=datos_existentes.get('medico', ''),
                    placeholder="Ej: Dr. Ana López"
                )
                fecha_nacimiento = st.date_input(
                    "Fecha de Nacimiento", 
                    value=datetime.strptime(datos_existentes.get('fecha_nacimiento', '01-01-1990'), "%d/%m/%Y") 
                    if datos_existentes.get('fecha_nacimiento') 
                    else None
                )
            
            # Información adicional
            st.markdown("### Información Adicional")
            col3, col4 = st.columns(2)
            with col3:
                telefono = st.text_input(
                    "Teléfono", 
                    value=datos_existentes.get('telefono', ''),
                    placeholder="+34 600 000 000"
                )
                email = st.text_input(
                    "Email", 
                    value=datos_existentes.get('email', ''),
                    placeholder="paciente@email.com"
                )
            with col4:
                direccion = st.text_area(
                    "Dirección", 
                    value=datos_existentes.get('direccion', ''),
                    placeholder="Calle..."
                )
                observaciones = st.text_area(
                    "Observaciones clínicas", 
                    value=datos_existentes.get('observaciones', ''),
                    placeholder="Observaciones relevantes..."
                )
            
            submitted = st.form_submit_button("💾 Guardar Datos del Paciente")
            
            if submitted:
                # Guardar todos los datos (incluyendo campos vacíos)
                fecha_nacimiento_str = fecha_nacimiento.strftime("%d/%m/%Y") if fecha_nacimiento else ""
                
                st.session_state.pacientes_data[paciente_seleccionado] = {
                    'nombre': nombre,
                    'edad': edad,
                    'sexo': sexo,
                    'id_paciente': id_paciente,
                    'medico': medico,
                    'fecha_nacimiento': fecha_nacimiento_str,
                    'telefono': telefono,
                    'email': email,
                    'direccion': direccion,
                    'observaciones': observaciones
                }
                st.success(f"✅ Datos de {paciente_seleccionado} guardados correctamente!")
        
        # Mostrar resumen de pacientes con datos guardados
        st.markdown("---")
        st.markdown("### 📋 Resumen de Pacientes")
        
        pacientes_guardados = st.session_state.pacientes_data.keys()
        if pacientes_guardados:
            st.info(f"**Pacientes con datos guardados:** {len(pacientes_guardados)}/{len(pacientes_disponibles)}")
            
            # Crear una tabla resumen
            resumen_data = []
            for paciente in pacientes_disponibles:
                tiene_datos = paciente in st.session_state.pacientes_data
                datos = st.session_state.pacientes_data.get(paciente, {})
                resumen_data.append({
                    "Paciente": paciente,
                    "Nombre": datos.get('nombre', 'No completado'),
                    "Edad": datos.get('edad', 'No completado'),
                    "Estado": "✅ Completado" if tiene_datos else "⏳ Pendiente"
                })
            
            # Mostrar tabla resumen
            df_resumen = pd.DataFrame(resumen_data)
            st.dataframe(
                df_resumen, 
                use_container_width=True,
                column_config={
                    "Paciente": "Código Paciente",
                    "Nombre": "Nombre",
                    "Edad": "Edad", 
                    "Estado": "Estado"
                }
            )
        else:
            st.warning("ℹ️ No hay datos de pacientes guardados todavía. Completa el formulario para guardar la información.")
        
    # Sección 3: Generar Reporte
    elif page == "📄 Generar Reporte":
        st.markdown('<div class="sub-header">📄 GENERAR REPORTES PDF</div>', unsafe_allow_html=True)
        
        pacientes_disponibles = list(st.session_state.resultado.keys())
        
        # Selector múltiple para generar varios PDFs a la vez
        st.markdown("### Selecciona los pacientes para generar reportes:")
        
        pacientes_seleccionados = st.multiselect(
            "Pacientes a generar:",
            pacientes_disponibles,
            default=list(st.session_state.pacientes_data.keys())  # Por defecto los que tienen datos
        )
        
        if not pacientes_seleccionados:
            st.info("💡 Selecciona al menos un paciente para generar reportes PDF")
            return
        
        # Opción para generar todos los pacientes
        if st.checkbox("Seleccionar todos los pacientes"):
            pacientes_seleccionados = pacientes_disponibles
        
        # Generar PDFs
        if st.button("🖨️ Generar Reportes PDF Seleccionados"):
            for paciente in pacientes_seleccionados:
                with st.spinner(f"Generando reporte para {paciente}..."):
                    try:
                        # Obtener datos del paciente (si existen)
                        paciente_data = st.session_state.pacientes_data.get(paciente, {
                            'nombre': '',
                            'edad': '',
                            'sexo': '',
                            'id_paciente': paciente,
                            'medico': '',
                            'fecha_nacimiento': '',
                            'telefono': '',
                            'email': '',
                            'direccion': '',
                            'observaciones': ''
                        })
                        
                        # Obtener datos genéticos del resultado_final
                        datos_geneticos = st.session_state.resultado.get(paciente, {})
                        
                        pdf = create_pdf(paciente_data, datos_geneticos, paciente)
                        
                        # Guardar PDF en bytes
                        pdf_bytes = pdf.output(dest='S').encode('latin1')
                        
                        # Nombre del archivo
                        nombre_paciente = paciente_data.get('nombre', paciente).replace(' ', '_')
                        nombre_archivo = f"reporte_{nombre_paciente}_{datetime.now().strftime('%Y%m%d')}.pdf"
                        
                        # Botón de descarga individual
                        col1, col2 = st.columns([3, 1])
                        with col1:
                            st.download_button(
                                label=f"📥 Descargar {paciente} - {paciente_data.get('nombre', 'Sin nombre')}",
                                data=pdf_bytes,
                                file_name=nombre_archivo,
                                mime="application/pdf",
                                key=f"download_{paciente}"
                            )
                        with col2:
                            st.info(f"Genes: {len(datos_geneticos)}")
                        
                    except Exception as e:
                        st.error(f"❌ Error al generar PDF para {paciente}: {str(e)}")
            
            st.success(f"✅ Se generaron {len(pacientes_seleccionados)} reportes correctamente!")

if __name__ == "__main__":
    main()



